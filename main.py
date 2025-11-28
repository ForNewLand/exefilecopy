import tkinter as tk
from tkinter import ttk, filedialog, messagebox, scrolledtext
import openpyxl
from openpyxl.utils import column_index_from_string
import threading
import os
import sys


class ExcelMatcherGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Excel 数据匹配与传输工具")
        self.root.geometry("900x750")

        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=tk.BOTH, expand=True)

        self.create_widgets()
        self.create_log_area()

    def create_widgets(self):
        # 标题
        title_label = ttk.Label(self.main_frame, text="Excel 数据匹配与传输工具",
                                font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=6, pady=(0, 20))

        # 文件选择区域
        file_frame = ttk.LabelFrame(self.main_frame, text="文件选择", padding="10")
        file_frame.grid(row=1, column=0, columnspan=6, sticky=(tk.W, tk.E), pady=(0, 10))
        file_frame.columnconfigure(1, weight=1)

        # 表格1文件选择
        ttk.Label(file_frame, text="表格1 (源文件):").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.table1_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.table1_path).grid(row=0, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(file_frame, text="浏览", command=self.browse_table1).grid(row=0, column=2, padx=5)

        # 表格2文件选择
        ttk.Label(file_frame, text="表格2 (目标文件):").grid(row=1, column=0, sticky=tk.W, pady=5)
        self.table2_path = tk.StringVar()
        ttk.Entry(file_frame, textvariable=self.table2_path).grid(row=1, column=1, sticky=(tk.W, tk.E), padx=5)
        ttk.Button(file_frame, text="浏览", command=self.browse_table2).grid(row=1, column=2, padx=5)

        # 配置区域 - 表格1
        config_frame1 = ttk.LabelFrame(self.main_frame, text="表格1配置 (源数据)", padding="10")
        config_frame1.grid(row=2, column=0, columnspan=6, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Label(config_frame1, text="工作表名称:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.table1_sheet = tk.StringVar(value="出勤")
        ttk.Entry(config_frame1, textvariable=self.table1_sheet, width=15).grid(row=0, column=1, sticky=tk.W, padx=5)

        ttk.Label(config_frame1, text="匹配列 (字母):").grid(row=0, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        self.table1_match_col = tk.StringVar(value="B")
        ttk.Entry(config_frame1, textvariable=self.table1_match_col, width=5).grid(row=0, column=3, sticky=tk.W, padx=5)

        ttk.Label(config_frame1, text="起始行号:").grid(row=0, column=4, sticky=tk.W, pady=5, padx=(20, 0))
        self.table1_start_row = tk.StringVar(value="5")
        ttk.Entry(config_frame1, textvariable=self.table1_start_row, width=5).grid(row=0, column=5, sticky=tk.W, padx=5)

        # 配置区域 - 表格2
        config_frame2 = ttk.LabelFrame(self.main_frame, text="表格2配置 (目标数据)", padding="10")
        config_frame2.grid(row=3, column=0, columnspan=6, sticky=(tk.W, tk.E), pady=(0, 10))

        ttk.Label(config_frame2, text="工作表名称:").grid(row=0, column=0, sticky=tk.W, pady=5)
        self.table2_sheet = tk.StringVar(value="班级1")
        ttk.Entry(config_frame2, textvariable=self.table2_sheet, width=15).grid(row=0, column=1, sticky=tk.W, padx=5)

        ttk.Label(config_frame2, text="匹配列 (字母):").grid(row=0, column=2, sticky=tk.W, pady=5, padx=(20, 0))
        self.table2_match_col = tk.StringVar(value="B")
        ttk.Entry(config_frame2, textvariable=self.table2_match_col, width=5).grid(row=0, column=3, sticky=tk.W, padx=5)

        # 多列映射区域 - 横向布局
        mapping_frame = ttk.LabelFrame(self.main_frame, text="列映射设置 (最多4列)", padding="10")
        mapping_frame.grid(row=4, column=0, columnspan=6, sticky=(tk.W, tk.E), pady=(0, 10))

        # 表头
        ttk.Label(mapping_frame, text="表格1 → 表格2", font=("Arial", 10, "bold")).grid(row=0, column=0, columnspan=2,
                                                                                        pady=(0, 10))
        ttk.Label(mapping_frame, text="源列", font=("Arial", 9, "bold")).grid(row=1, column=0, padx=10)
        ttk.Label(mapping_frame, text="目标列", font=("Arial", 9, "bold")).grid(row=1, column=1, padx=10)

        # 创建4对列映射输入框
        self.table1_value_cols = []
        self.table2_target_cols = []

        for i in range(4):
            # 表格1源列
            value_col_var = tk.StringVar()
            value_col_entry = ttk.Entry(mapping_frame, textvariable=value_col_var, width=5)
            value_col_entry.grid(row=2 + i, column=0, padx=10, pady=5)
            self.table1_value_cols.append(value_col_var)

            # 箭头
            ttk.Label(mapping_frame, text="→").grid(row=2 + i, column=1, padx=5)

            # 表格2目标列
            target_col_var = tk.StringVar()
            target_col_entry = ttk.Entry(mapping_frame, textvariable=target_col_var, width=5)
            target_col_entry.grid(row=2 + i, column=2, padx=10, pady=5)
            self.table2_target_cols.append(target_col_var)

            # 列标签
            ttk.Label(mapping_frame, text=f"列 {i + 1}").grid(row=2 + i, column=3, padx=10, pady=5, sticky=tk.W)

        # 设置默认值
        self.table1_value_cols[0].set("P")
        self.table2_target_cols[0].set("D")

        # 按钮区域
        button_frame = ttk.Frame(self.main_frame)
        button_frame.grid(row=5, column=0, columnspan=6, pady=20)

        ttk.Button(button_frame, text="开始匹配", command=self.start_matching).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="清除日志", command=self.clear_log).pack(side=tk.LEFT, padx=10)
        ttk.Button(button_frame, text="退出", command=self.root.quit).pack(side=tk.LEFT, padx=10)

        # 进度条
        self.progress = ttk.Progressbar(self.main_frame, mode='indeterminate')
        self.progress.grid(row=6, column=0, columnspan=6, sticky=(tk.W, tk.E), pady=10)

    def create_log_area(self):
        # 日志区域
        log_frame = ttk.LabelFrame(self.main_frame, text="处理日志", padding="10")
        log_frame.grid(row=7, column=0, columnspan=6, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(0, 10))
        self.main_frame.rowconfigure(7, weight=1)

        self.log_area = scrolledtext.ScrolledText(log_frame, width=100, height=15)
        self.log_area.pack(fill=tk.BOTH, expand=True)

    def browse_table1(self):
        filename = filedialog.askopenfilename(
            title="选择表格1文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.table1_path.set(filename)

    def browse_table2(self):
        filename = filedialog.askopenfilename(
            title="选择表格2文件",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All files", "*.*")]
        )
        if filename:
            self.table2_path.set(filename)

    def log(self, message):
        self.log_area.insert(tk.END, message + "\n")
        self.log_area.see(tk.END)
        self.root.update_idletasks()

    def clear_log(self):
        self.log_area.delete(1.0, tk.END)

    def validate_inputs(self):
        # 检查文件路径
        if not self.table1_path.get():
            messagebox.showerror("错误", "请选择表格1文件")
            return False

        if not self.table2_path.get():
            messagebox.showerror("错误", "请选择表格2文件")
            return False

        # 检查文件是否存在
        if not os.path.exists(self.table1_path.get()):
            messagebox.showerror("错误", f"表格1文件不存在: {self.table1_path.get()}")
            return False

        if not os.path.exists(self.table2_path.get()):
            messagebox.showerror("错误", f"表格2文件不存在: {self.table2_path.get()}")
            return False

        # 检查匹配列字母格式
        try:
            column_index_from_string(self.table1_match_col.get().upper())
            column_index_from_string(self.table2_match_col.get().upper())
        except:
            messagebox.showerror("错误", "匹配列字母格式不正确，请使用A、B、C等格式")
            return False

        # 检查映射列字母格式
        for i in range(4):
            value_col = self.table1_value_cols[i].get().strip().upper()
            target_col = self.table2_target_cols[i].get().strip().upper()

            # 如果一列有值而另一列没有，提示错误
            if (value_col and not target_col) or (not value_col and target_col):
                messagebox.showerror("错误", f"列 {i + 1} 的源列和目标列必须同时填写或同时为空")
                return False

            # 如果两列都有值，检查格式
            if value_col and target_col:
                try:
                    column_index_from_string(value_col)
                    column_index_from_string(target_col)
                except:
                    messagebox.showerror("错误", f"列 {i + 1} 的列字母格式不正确，请使用A、B、C等格式")
                    return False

        # 检查是否至少有一对映射列
        has_mapping = False
        for i in range(4):
            if self.table1_value_cols[i].get().strip() and self.table2_target_cols[i].get().strip():
                has_mapping = True
                break

        if not has_mapping:
            messagebox.showerror("错误", "请至少设置一对列映射")
            return False

        # 检查行号
        try:
            start_row = int(self.table1_start_row.get())
            if start_row < 1:
                raise ValueError("行号必须大于0")
        except ValueError:
            messagebox.showerror("错误", "起始行号必须是大于0的整数")
            return False

        return True

    def start_matching(self):
        if not self.validate_inputs():
            return

        # 在新线程中执行匹配操作，避免GUI冻结
        thread = threading.Thread(target=self.execute_matching)
        thread.daemon = True
        thread.start()

    def execute_matching(self):
        try:
            # 开始进度条
            self.progress.start()
            self.log("开始处理数据...")

            # 获取配置
            config = {
                'table1_path': self.table1_path.get(),
                'table2_path': self.table2_path.get(),
                'table1_sheet': self.table1_sheet.get(),
                'table2_sheet': self.table2_sheet.get(),
                'table1_match_col': self.table1_match_col.get().upper(),
                'table2_match_col': self.table2_match_col.get().upper(),
                'table1_start_row': int(self.table1_start_row.get())
            }

            # 获取列映射
            column_mappings = []
            for i in range(4):
                value_col = self.table1_value_cols[i].get().strip().upper()
                target_col = self.table2_target_cols[i].get().strip().upper()
                if value_col and target_col:
                    column_mappings.append((value_col, target_col))

            self.log(f"已设置 {len(column_mappings)} 对列映射")
            for i, (value_col, target_col) in enumerate(column_mappings):
                self.log(f"  映射 {i + 1}: 表格1.{value_col} → 表格2.{target_col}")

            # 加载工作簿
            self.log(f"加载表格1: {config['table1_path']}")
            wb1 = openpyxl.load_workbook(config['table1_path'], data_only=True)

            self.log(f"加载表格2: {config['table2_path']}")
            wb2 = openpyxl.load_workbook(config['table2_path'])

            # 获取工作表
            sheet1 = wb1[config['table1_sheet']]
            sheet2 = wb2[config['table2_sheet']]

            # 将列字母转换为数字索引
            table1_match_col_idx = column_index_from_string(config['table1_match_col'])
            table2_match_col_idx = column_index_from_string(config['table2_match_col'])

            # 构建列映射的索引
            value_col_indices = []
            target_col_indices = []
            for value_col, target_col in column_mappings:
                value_col_indices.append(column_index_from_string(value_col))
                target_col_indices.append(column_index_from_string(target_col))

            # 构建表格2的匹配索引
            self.log("正在构建匹配索引...")
            name_to_row = {}

            for row_num in range(1, sheet2.max_row + 1):
                match_cell = sheet2.cell(row=row_num, column=table2_match_col_idx)
                if match_cell.value:
                    name = str(match_cell.value).strip()
                    if name:
                        name_to_row[name] = row_num

            self.log(f"表格2中共找到 {len(name_to_row)} 个匹配项")

            # 处理表格1的数据
            processed_count = 0
            matched_count = 0
            not_matched_count = 0

            self.log(f"\n开始处理表格1数据...")

            for row_num in range(config['table1_start_row'], sheet1.max_row + 1):
                match_cell = sheet1.cell(row=row_num, column=table1_match_col_idx)

                # 跳过空单元格
                if not match_cell.value:
                    continue

                name = str(match_cell.value).strip()
                processed_count += 1

                # 在表格2中查找匹配
                if name in name_to_row:
                    matched_row = name_to_row[name]

                    # 复制所有映射列的值
                    copied_values = []
                    for i, (value_col_idx, target_col_idx) in enumerate(zip(value_col_indices, target_col_indices)):
                        value_cell = sheet1.cell(row=row_num, column=value_col_idx)
                        value = value_cell.value

                        # 写入表格2的目标列
                        sheet2.cell(row=matched_row, column=target_col_idx).value = value
                        copied_values.append(f"{column_mappings[i][0]}→{column_mappings[i][1]}: {value}")

                    matched_count += 1
                    self.log(f"✓ 匹配成功: '{name}' -> 表格2第{matched_row}行")
                    for value_info in copied_values:
                        self.log(f"    写入: {value_info}")
                else:
                    not_matched_count += 1
                    match_cell_ref = f"{config['table1_match_col']}{row_num}"
                    self.log(f"✗ 未找到匹配: '{name}' (位置: {match_cell_ref})")

            # 保存结果
            self.log(f"\n正在保存结果...")
            wb2.save(config['table2_path'])

            # 输出统计信息
            self.log("\n" + "=" * 50)
            self.log("处理完成！统计信息：")
            self.log(f"总共处理: {processed_count} 个项目")
            self.log(f"成功匹配: {matched_count} 个")
            self.log(f"未找到匹配: {not_matched_count} 个")
            self.log(f"列映射数量: {len(column_mappings)} 对")
            self.log(f"结果文件: {config['table2_path']}")
            self.log("=" * 50)

            # 关闭工作簿
            wb1.close()
            wb2.close()

            # 显示完成消息
            self.root.after(0, lambda: messagebox.showinfo("完成",
                                                           f"处理完成！\n总共处理: {processed_count} 个项目\n"
                                                           f"成功匹配: {matched_count} 个\n"
                                                           f"未找到匹配: {not_matched_count} 个\n"
                                                           f"列映射数量: {len(column_mappings)} 对"))

        except FileNotFoundError as e:
            error_msg = f"❌ 错误: 文件未找到 - {e}"
            self.log(error_msg)
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        except KeyError as e:
            error_msg = f"❌ 错误: 工作表 '{e}' 不存在"
            self.log(error_msg)
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        except PermissionError as e:
            error_msg = f"❌ 错误: 文件被占用或无写入权限 - {e}"
            self.log(error_msg)
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        except Exception as e:
            error_msg = f"❌ 错误: 处理过程中出现未知错误 - {e}"
            self.log(error_msg)
            self.root.after(0, lambda: messagebox.showerror("错误", error_msg))
        finally:
            # 停止进度条
            self.progress.stop()

    def run(self):
        self.root.mainloop()


def create_executable():
    """创建可执行文件的函数"""
    try:
        import PyInstaller.__main__

        # 将当前脚本保存为单独的文件
        script_content = """
# 这里是上面完整的Python代码
"""

        # 写入临时脚本文件
        with open("excel_matcher_gui.py", "w", encoding="utf-8") as f:
            f.write(script_content)

        # 使用PyInstaller创建可执行文件
        PyInstaller.__main__.run([
            'excel_matcher_gui.py',
            '--onefile',
            '--windowed',
            '--name=Excel数据匹配工具',
            '--icon=NONE',  # 如果有图标文件可以替换
        ])

        print("可执行文件创建成功！")
    except ImportError:
        print("请先安装PyInstaller: pip install pyinstaller")
    except Exception as e:
        print(f"创建可执行文件时出错: {e}")


if __name__ == "__main__":
    # 检查是否要创建可执行文件
    if len(sys.argv) > 1 and sys.argv[1] == "--build":
        create_executable()
    else:
        root = tk.Tk()
        app = ExcelMatcherGUI(root)
        app.run()